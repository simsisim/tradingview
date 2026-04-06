"""
generate_SC_xlsx.py
Creates SC_sector_watchlists.xlsx from StockCharts sector drill-down HTML files.

Sheets : one per sector (named SC_com_ser, SC_con_dis, ...)
Columns: Ticker | Market Cap | Industry | Sector
"""

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
FULLWIDTH_COLON = "\uff1a"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


# ── naming helpers (same logic as generate_SC_watchlists.py) ─────────────────

def abbrev(name: str, n: int = 3) -> str:
    words = re.split(r"[\s_]+", name.strip())
    return "_".join(w[:n].lower() for w in words if w)


def resolve_abbrevs(suffixes: list) -> dict:
    result = {s: abbrev(s) for s in suffixes}
    changed = True
    while changed:
        changed = False
        by_abbrev = {}
        for suffix, ab in result.items():
            by_abbrev.setdefault(ab, []).append(suffix)
        for ab, suffixes_list in by_abbrev.items():
            if len(suffixes_list) > 1:
                cur_n = len(ab.split("_")[0])
                new_n = cur_n + 1
                for s in suffixes_list:
                    result[s] = abbrev(s, new_n)
                changed = True
    return result


def sector_sheet_name(sector_name: str) -> str:
    return f"SC_{abbrev(sector_name)}"


# ── HTML parsing ──────────────────────────────────────────────────────────────

def extract_timestamp(fname: str):
    m = re.search(
        r"\((\d+)_(\d+)_(\d+)\s+(\d+)" + FULLWIDTH_COLON +
        r"(\d+)" + FULLWIDTH_COLON + r"(\d+)\s+(AM|PM)\)",
        fname,
    )
    if not m:
        return (0, 0, 0, 0, 0, 0)
    month, day, year, hour, minute, second, ampm = (
        int(m.group(1)), int(m.group(2)), int(m.group(3)),
        int(m.group(4)), int(m.group(5)), int(m.group(6)),
        m.group(7),
    )
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    return (year, month, day, hour, minute, second)


def get_industry_suffix(fname: str) -> str:
    m = re.search(r"PM\)_(.+?)\.html$", fname)
    return m.group(1) if m else re.sub(r"\.html$", "", fname)


MKTCAP_RE = re.compile(r"^[\d,]+\.?\d*\s+[BMT]$")


NAME_RE = re.compile(r'<a href="https://stockcharts\.com/sc3/ui\?s=[^"]+">([^<]+)</a>')


def parse_stocks(html_path: str) -> list[tuple[str, str, str]]:
    """Return list of (ticker, name, market_cap) from one HTML file."""
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    results = []
    for m in re.finditer(
        r'<span class=symlink data-sym>([A-Z][A-Z0-9.]{0,6})</span>', content
    ):
        ticker = m.group(1)
        chunk = content[m.end(): m.end() + 800]
        stop = chunk.find("<tr ")
        chunk = chunk[: stop] if stop != -1 else chunk

        # Company name: first <a> href pointing to sc3/ui
        name_m = NAME_RE.search(chunk)
        name = name_m.group(1).strip() if name_m else ""

        td_vals = re.findall(r"<td[^>]*>([^<]{1,30})", chunk)
        mktcap = ""
        for v in td_vals:
            v = v.strip()
            if MKTCAP_RE.match(v):
                mktcap = v
                break
        results.append((ticker, name, mktcap))

    return results


# ── sheet builder ─────────────────────────────────────────────────────────────

def write_sheet(ws, sector_name: str, sector_path: str):
    html_files = sorted(
        [f for f in os.listdir(sector_path) if f.endswith(".html")],
        key=extract_timestamp,
    )
    industries = [get_industry_suffix(f) for f in html_files]
    abbrev_map = resolve_abbrevs(industries)

    # Header
    ws.append(["Ticker", "Name", "Market Cap", "Industry", "Sector"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # Alternating row colours per industry block
    COLOURS = ["EBF3FB", "FFFFFF"]
    colour_idx = 0
    total = 0

    for fname, industry in zip(html_files, industries):
        section = abbrev_map[industry]
        stocks = parse_stocks(os.path.join(sector_path, fname))
        fill = PatternFill("solid", fgColor=COLOURS[colour_idx % 2])
        colour_idx += 1

        for ticker, name, mktcap in stocks:
            ws.append([ticker, name, mktcap, section, sector_name])
            for cell in ws[ws.max_row]:
                cell.fill = fill
            total += 1

        print(f"  {industry:40} → ### {section:18}  ({len(stocks)} stocks)")

    # Column widths
    col_widths = [10, 32, 14, 20, 26]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    print(f"  → sheet '{ws.title}'  [{total} rows]\n")
    return total


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    sectors = [
        (name, os.path.join(BASE_DIR, name))
        for name in os.listdir(BASE_DIR)
        if os.path.isdir(os.path.join(BASE_DIR, name))
        and not name.startswith(".")
        and name != "output"
    ]

    def sort_key(item):
        return (0 if item[0] == "Communication Services" else 1, item[0])

    sectors.sort(key=sort_key)

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    grand_total = 0
    for sector_name, sector_path in sectors:
        sheet_name = sector_sheet_name(sector_name)
        print(f"=== {sector_name} → sheet '{sheet_name}' ===")
        ws = wb.create_sheet(title=sheet_name)
        grand_total += write_sheet(ws, sector_name, sector_path)

    out_path = os.path.join(OUTPUT_DIR, "SC_sector_watchlists.xlsx")
    wb.save(out_path)
    print(f"{'='*60}")
    print(f"Saved: {out_path}")
    print(f"GRAND TOTAL: {grand_total} rows across {len(sectors)} sheets")


if __name__ == "__main__":
    main()
